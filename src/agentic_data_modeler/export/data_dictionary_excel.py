"""Structural Source Data Dictionary -> Excel (Phase-9 review/consumption export).

Deterministic and LLM-free: it renders the OBSERVED facts captured by Phase 1
(source_object_observation + source_attribute_observation) into an .xlsx
workbook. It contains no business meaning — those columns are produced later by
the semantic phase. The workbook is a review surface generated from the
authoritative records; it is never re-imported as a source of truth.
"""

from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=14)
_NOTE_FONT = Font(italic=True, color="9C5700")


def _style_header(ws: Worksheet, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}1"


def _autosize(ws: Worksheet, widths: list[int]) -> None:
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _constraint_summary(obj: dict[str, Any]) -> str:
    parts = []
    for c in obj.get("constraint_observations", []) or []:
        ctype = c.get("constraint_type", "")
        cols = ""
        details = c.get("constraint_details")
        if details:
            try:
                cols = ",".join(json.loads(details).get("columns", []))
            except (json.JSONDecodeError, AttributeError):
                cols = ""
        parts.append(f"{ctype}({cols})" if cols else ctype)
    return "; ".join(parts)


def build_source_dictionary_workbook(
    objects: list[dict[str, Any]],
    attributes: list[dict[str, Any]],
    *,
    meta: dict[str, Any],
    out_path: str | Path,
    example: bool = False,
) -> Path:
    """Write a structural Source Data Dictionary workbook. Returns the path."""
    wb = Workbook()

    # --- Cover ---
    cover = wb.active
    cover.title = "Cover"
    cover["A1"] = "Source Data Dictionary — Structural"
    cover["A1"].font = _TITLE_FONT
    rows = [
        ("Solution run", meta.get("run_id", "")),
        ("Source catalog", meta.get("catalog", "")),
        ("Source schema", meta.get("schema", "")),
        ("Scope mode", meta.get("scope_mode", "")),
        ("Source snapshot", meta.get("source_snapshot_id", "")),
        ("Objects", len(objects)),
        ("Attributes", len(attributes)),
        ("Generated at", meta.get("generated_at") or datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%SZ")),
        ("Content", "Structural facts only (OBSERVED). Business meaning pending the semantic phase."),
    ]
    for i, (label, value) in enumerate(rows, start=3):
        cover.cell(row=i, column=1, value=label).font = Font(bold=True)
        cover.cell(row=i, column=2, value=value)
    if example:
        note = cover.cell(
            row=3 + len(rows) + 1, column=1,
            value="EXAMPLE / FORMAT DEMO — rows are a unit-test fixture, not real source "
                  "metadata. The real workbook is generated from Phase-1 evidence on Databricks.")
        note.font = _NOTE_FONT
    _autosize(cover, [18, 90])

    # --- Objects ---
    ws_o = wb.create_sheet("Objects")
    o_headers = ["Catalog", "Schema", "Object", "Type", "Attributes", "Keys / Constraints"]
    ws_o.append(o_headers)
    for obj in sorted(objects, key=lambda o: o.get("object_name", "")):
        ws_o.append([
            obj.get("catalog_name", ""), obj.get("schema_name", ""), obj.get("object_name", ""),
            obj.get("object_type", ""), obj.get("attribute_count", ""), _constraint_summary(obj),
        ])
    _style_header(ws_o, len(o_headers))
    _autosize(ws_o, [22, 18, 28, 10, 12, 40])

    # --- Attributes ---
    ws_a = wb.create_sheet("Attributes")
    a_headers = ["Object", "#", "Attribute", "Data Type", "Nullable", "Key Role",
                 "Default", "Length", "Precision", "Scale"]
    ws_a.append(a_headers)
    for attr in sorted(attributes, key=lambda a: (a.get("object_name", ""), a.get("ordinal_position", 0))):
        ws_a.append([
            attr.get("object_name", ""), attr.get("ordinal_position", ""), attr.get("attribute_name", ""),
            attr.get("data_type", ""), "YES" if attr.get("nullable") else "NO",
            attr.get("constraint_role", ""), attr.get("default_value", ""),
            attr.get("length", ""), attr.get("precision", ""), attr.get("scale", ""),
        ])
    _style_header(ws_a, len(a_headers))
    _autosize(ws_a, [26, 5, 28, 16, 9, 13, 16, 8, 10, 7])

    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


def build_full_source_dictionary_workbook(
    objects: list[dict[str, Any]],
    attributes: list[dict[str, Any]],
    dictionary_attributes: list[dict[str, Any]],
    code_values: list[dict[str, Any]],
    open_questions: list[dict[str, Any]],
    relationships: list[dict[str, Any]] | None = None,
    *,
    meta: dict[str, Any],
    out_path: str | Path,
) -> Path:
    """Write a full semantic Source Data Dictionary workbook (includes business names, definitions, code values, etc.). Returns the path."""
    wb = Workbook()

    # --- Cover ---
    cover = wb.active
    cover.title = "Cover"
    cover["A1"] = "Source Data Dictionary — Full Semantic"
    cover["A1"].font = _TITLE_FONT
    
    unresolved_count = sum(1 for d in dictionary_attributes if d.get("business_definition", {}).get("trust_state") == "UNRESOLVED")
    privacy_flagged_count = sum(1 for d in dictionary_attributes if d.get("privacy_classification") not in ("", None))
    
    rows = [
        ("Solution run", meta.get("run_id", "")),
        ("Source catalog", meta.get("catalog", "")),
        ("Source schema", meta.get("schema", "")),
        ("Scope mode", meta.get("scope_mode", "")),
        ("Source snapshot", meta.get("source_snapshot_id", "")),
        ("Context snapshot", meta.get("context_snapshot_id", "")),
        ("Objects", len(objects)),
        ("Attributes", len(attributes)),
        ("Dictionary attributes", len(dictionary_attributes)),
        ("UNRESOLVED", unresolved_count),
        ("Privacy-flagged", privacy_flagged_count),
        ("Code values", len(code_values)),
        ("Generated at", meta.get("generated_at") or datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%SZ")),
        ("Content", "Full semantic SDD with business names, definitions, trust state, and code value meanings."),
    ]
    for i, (label, value) in enumerate(rows, start=3):
        cover.cell(row=i, column=1, value=label).font = Font(bold=True)
        cover.cell(row=i, column=2, value=value)
    _autosize(cover, [20, 90])

    # --- Objects (reuse structural) ---
    ws_o = wb.create_sheet("Objects")
    o_headers = ["Catalog", "Schema", "Object", "Type", "Attributes", "Keys / Constraints"]
    ws_o.append(o_headers)
    for obj in sorted(objects, key=lambda o: o.get("object_name", "")):
        ws_o.append([
            obj.get("catalog_name", ""), obj.get("schema_name", ""), obj.get("object_name", ""),
            obj.get("object_type", ""), obj.get("attribute_count", ""), _constraint_summary(obj),
        ])
    _style_header(ws_o, len(o_headers))
    _autosize(ws_o, [22, 18, 28, 10, 12, 40])

    # --- Attributes (reuse structural) ---
    ws_a = wb.create_sheet("Attributes")
    a_headers = ["Object", "#", "Attribute", "Data Type", "Nullable", "Key Role",
                 "Default", "Length", "Precision", "Scale"]
    ws_a.append(a_headers)
    for attr in sorted(attributes, key=lambda a: (a.get("object_name", ""), a.get("ordinal_position", 0))):
        ws_a.append([
            attr.get("object_name", ""), attr.get("ordinal_position", ""), attr.get("attribute_name", ""),
            attr.get("data_type", ""), "YES" if attr.get("nullable") else "NO",
            attr.get("constraint_role", ""), attr.get("default_value", ""),
            attr.get("length", ""), attr.get("precision", ""), attr.get("scale", ""),
        ])
    _style_header(ws_a, len(a_headers))
    _autosize(ws_a, [26, 5, 28, 16, 9, 13, 16, 8, 10, 7])

    # --- Dictionary (semantic) ---
    ws_d = wb.create_sheet("Dictionary")
    d_headers = ["Object", "Attribute", "Business Name", "Business Definition", 
                 "Name Trust", "Definition Trust", "Confidence", "Evidence Count",
                 "Privacy", "Lifecycle State", "Review Decision"]
    ws_d.append(d_headers)
    for d in sorted(dictionary_attributes, key=lambda x: (x.get("object_name", ""), x.get("attribute_name", ""))):
        biz_name = d.get("business_name", {}) or {}
        biz_def = d.get("business_definition", {}) or {}
        review = d.get("review_decision", {}) or {}
        ws_d.append([
            d.get("object_name", ""),
            d.get("attribute_name", ""),
            biz_name.get("value", "") if biz_name.get("trust_state") != "UNRESOLVED" else "",
            biz_def.get("value", "") if biz_def.get("trust_state") != "UNRESOLVED" else "",
            biz_name.get("trust_state", ""),
            biz_def.get("trust_state", ""),
            d.get("confidence", ""),
            len(d.get("evidence_refs", []) or []),
            d.get("privacy_classification", ""),
            d.get("lifecycle_state", ""),
            review.get("decision", ""),
        ])
    _style_header(ws_d, len(d_headers))
    _autosize(ws_d, [26, 28, 30, 50, 15, 18, 12, 14, 12, 16, 16])

    # --- Code Values ---
    ws_c = wb.create_sheet("Code Values")
    c_headers = ["Object", "Attribute", "Code", "Meaning", "Trust", "Confidence"]
    ws_c.append(c_headers)
    for c in sorted(code_values, key=lambda x: (x.get("object_name", ""), x.get("attribute_name", ""), x.get("code_value", ""))):
        meaning = c.get("code_meaning", {}) or {}
        ws_c.append([
            c.get("object_name", ""),
            c.get("attribute_name", ""),
            c.get("code_value", ""),
            meaning.get("value", ""),
            meaning.get("trust_state", ""),
            c.get("confidence", ""),
        ])
    _style_header(ws_c, len(c_headers))
    _autosize(ws_c, [26, 28, 20, 50, 15, 12])

    # --- Open Questions ---
    ws_q = wb.create_sheet("Open Questions")
    q_headers = ["Type", "Object", "Attribute", "Question"]
    ws_q.append(q_headers)
    for q in sorted(open_questions, key=lambda x: (x.get("object_name", ""), x.get("attribute_name", ""))):
        ws_q.append([
            q.get("question_type", ""),
            q.get("object_name", ""),
            q.get("attribute_name", ""),
            q.get("question_text", ""),
        ])
    _style_header(ws_q, len(q_headers))
    _autosize(ws_q, [20, 26, 28, 60])

    # --- Relationships (optional) ---
    if relationships:
        ws_r = wb.create_sheet("Relationships")
        r_headers = ["From", "To", "Type", "Trust", "Confidence"]
        ws_r.append(r_headers)
        for r in sorted(relationships, key=lambda x: (x.get("from_qualified_name", ""), x.get("to_qualified_name", ""))):
            rel_type = r.get("relationship_type", {}) or {}
            ws_r.append([
                r.get("from_qualified_name", ""),
                r.get("to_qualified_name", ""),
                rel_type.get("value", ""),
                rel_type.get("trust_state", ""),
                r.get("confidence", ""),
            ])
        _style_header(ws_r, len(r_headers))
        _autosize(ws_r, [40, 40, 20, 15, 12])

    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out
