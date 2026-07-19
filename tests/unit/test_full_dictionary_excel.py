"""Tests for build_full_source_dictionary_workbook (semantic SDD Excel export).

Verifies the full semantic workbook generation according to the requirements in
run-and-validate-source-discovery SKILL.md §1.
"""

import tempfile
from pathlib import Path

import pytest
from openpyxl import load_workbook

from agentic_data_modeler.export.data_dictionary_excel import build_full_source_dictionary_workbook


@pytest.fixture
def sample_data():
    """Sample data matching the semantic SDD schema."""
    objects = [
        {"catalog_name": "test_catalog", "schema_name": "test_schema", "object_name": "table1",
         "object_type": "TABLE", "attribute_count": 2, "constraint_observations": []},
    ]
    attributes = [
        {"object_name": "table1", "ordinal_position": 1, "attribute_name": "id",
         "data_type": "INTEGER", "nullable": False, "constraint_role": "PRIMARY KEY"},
        {"object_name": "table1", "ordinal_position": 2, "attribute_name": "name",
         "data_type": "STRING", "nullable": True, "constraint_role": ""},
    ]
    dictionary_attributes = [
        {
            "object_name": "table1", "attribute_name": "id",
            "business_name": {"value": "Identifier", "trust_state": "INFERRED"},
            "business_definition": {"value": "Unique record identifier", "trust_state": "INFERRED"},
            "confidence": 0.95, "evidence_refs": ["ev_001"],
            "privacy_classification": "", "lifecycle_state": "DRAFT",
            "review_decision": {"decision": "PENDING"},
        },
        {
            "object_name": "table1", "attribute_name": "name",
            "business_name": {"value": "", "trust_state": "UNRESOLVED"},
            "business_definition": {"value": "", "trust_state": "UNRESOLVED"},
            "confidence": 0.0, "evidence_refs": [],
            "privacy_classification": "PII", "lifecycle_state": "DRAFT",
            "review_decision": {"decision": "PENDING"},
        },
    ]
    code_values = [
        {
            "object_name": "table1", "attribute_name": "status", "code_value": "A",
            "code_meaning": {"value": "Active", "trust_state": "INFERRED"},
            "confidence": 0.9,
        },
    ]
    open_questions = [
        {
            "question_type": "BUSINESS_NAME", "object_name": "table1", "attribute_name": "name",
            "question_text": "What is the business meaning of 'name'?",
        },
    ]
    relationships = [
        {
            "from_qualified_name": "table1.id", "to_qualified_name": "table2.table1_id",
            "relationship_type": {"value": "FOREIGN_KEY", "trust_state": "OBSERVED"},
            "confidence": 1.0,
        },
    ]
    meta = {
        "run_id": "test_run_001",
        "catalog": "test_catalog",
        "schema": "test_schema",
        "scope_mode": "schema",
        "source_snapshot_id": "snap_001",
        "context_snapshot_id": "ctx_001",
    }
    return {
        "objects": objects,
        "attributes": attributes,
        "dictionary_attributes": dictionary_attributes,
        "code_values": code_values,
        "open_questions": open_questions,
        "relationships": relationships,
        "meta": meta,
    }


def test_sheet_names_and_order(sample_data):
    """Sheet names appear in the correct order."""
    with tempfile.TemporaryDirectory() as tmpdir:
        out_path = Path(tmpdir) / "test.xlsx"
        build_full_source_dictionary_workbook(
            sample_data["objects"],
            sample_data["attributes"],
            sample_data["dictionary_attributes"],
            sample_data["code_values"],
            sample_data["open_questions"],
            sample_data["relationships"],
            meta=sample_data["meta"],
            out_path=out_path,
        )
        wb = load_workbook(out_path)
        expected = ["Cover", "Objects", "Attributes", "Dictionary", "Code Values", "Open Questions", "Relationships"]
        assert wb.sheetnames == expected


def test_dictionary_coverage(sample_data):
    """Dictionary row count matches number of dictionary_attributes (100% coverage)."""
    with tempfile.TemporaryDirectory() as tmpdir:
        out_path = Path(tmpdir) / "test.xlsx"
        build_full_source_dictionary_workbook(
            sample_data["objects"],
            sample_data["attributes"],
            sample_data["dictionary_attributes"],
            sample_data["code_values"],
            sample_data["open_questions"],
            sample_data["relationships"],
            meta=sample_data["meta"],
            out_path=out_path,
        )
        wb = load_workbook(out_path)
        ws = wb["Dictionary"]
        # Subtract 1 for header row
        dict_rows = sum(1 for _ in ws.iter_rows(min_row=2, max_row=ws.max_row))
        assert dict_rows == len(sample_data["dictionary_attributes"])


def test_code_values_count(sample_data):
    """Code Values row count matches code_value records."""
    with tempfile.TemporaryDirectory() as tmpdir:
        out_path = Path(tmpdir) / "test.xlsx"
        build_full_source_dictionary_workbook(
            sample_data["objects"],
            sample_data["attributes"],
            sample_data["dictionary_attributes"],
            sample_data["code_values"],
            sample_data["open_questions"],
            sample_data["relationships"],
            meta=sample_data["meta"],
            out_path=out_path,
        )
        wb = load_workbook(out_path)
        ws = wb["Code Values"]
        code_rows = sum(1 for _ in ws.iter_rows(min_row=2, max_row=ws.max_row))
        assert code_rows == len(sample_data["code_values"])


def test_unresolved_shows_no_invented_value(sample_data):
    """UNRESOLVED attributes show blank values, not fabricated content."""
    with tempfile.TemporaryDirectory() as tmpdir:
        out_path = Path(tmpdir) / "test.xlsx"
        build_full_source_dictionary_workbook(
            sample_data["objects"],
            sample_data["attributes"],
            sample_data["dictionary_attributes"],
            sample_data["code_values"],
            sample_data["open_questions"],
            sample_data["relationships"],
            meta=sample_data["meta"],
            out_path=out_path,
        )
        wb = load_workbook(out_path)
        ws = wb["Dictionary"]
        # Row 3 is the UNRESOLVED 'name' attribute (header is row 1, first data is row 2)
        # Business Name (col 3) and Business Definition (col 4) should be empty for UNRESOLVED
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            name_trust = row[4].value  # Name Trust column
            def_trust = row[5].value    # Definition Trust column
            if name_trust == "UNRESOLVED":
                assert row[2].value == "" or row[2].value is None  # Business Name should be blank
            if def_trust == "UNRESOLVED":
                assert row[3].value == "" or row[3].value is None  # Business Definition should be blank


def test_no_fabricated_evidence_ids(sample_data):
    """No fabricated evidence IDs appear in any cell (evidence_refs are not rendered)."""
    with tempfile.TemporaryDirectory() as tmpdir:
        out_path = Path(tmpdir) / "test.xlsx"
        build_full_source_dictionary_workbook(
            sample_data["objects"],
            sample_data["attributes"],
            sample_data["dictionary_attributes"],
            sample_data["code_values"],
            sample_data["open_questions"],
            sample_data["relationships"],
            meta=sample_data["meta"],
            out_path=out_path,
        )
        wb = load_workbook(out_path)
        # Check all sheets for fabricated patterns like "ev_fake_", "evidence_fake_", etc.
        fabricated_patterns = ["fake", "fabricated", "invented"]
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value:
                        cell_str = str(cell.value).lower()
                        for pattern in fabricated_patterns:
                            assert pattern not in cell_str, f"Fabricated pattern '{pattern}' found in {sheet_name}"


def test_workbook_without_relationships(sample_data):
    """Workbook can be generated without the optional Relationships sheet."""
    with tempfile.TemporaryDirectory() as tmpdir:
        out_path = Path(tmpdir) / "test.xlsx"
        build_full_source_dictionary_workbook(
            sample_data["objects"],
            sample_data["attributes"],
            sample_data["dictionary_attributes"],
            sample_data["code_values"],
            sample_data["open_questions"],
            relationships=None,  # No relationships
            meta=sample_data["meta"],
            out_path=out_path,
        )
        wb = load_workbook(out_path)
        expected = ["Cover", "Objects", "Attributes", "Dictionary", "Code Values", "Open Questions"]
        assert wb.sheetnames == expected
