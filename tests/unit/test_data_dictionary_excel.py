"""Unit test for the structural Source Data Dictionary Excel exporter."""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "src"))

from agentic_data_modeler.export import build_source_dictionary_workbook

OBJECTS = [
    {"catalog_name": "insurance_source_discovery", "schema_name": "gw_pc_bronze",
     "object_name": "pc_policy", "object_type": "TABLE", "attribute_count": 2,
     "constraint_observations": [
         {"constraint_type": "PRIMARY KEY", "constraint_details": '{"columns":["pol_id"]}'}]},
]
ATTRIBUTES = [
    {"object_name": "pc_policy", "attribute_name": "pol_id", "ordinal_position": 1,
     "data_type": "STRING", "nullable": False, "constraint_role": "PRIMARY_KEY"},
    {"object_name": "pc_policy", "attribute_name": "wrtn_prem_amt", "ordinal_position": 2,
     "data_type": "DECIMAL(12,2)", "nullable": True, "constraint_role": "NONE",
     "precision": 12, "scale": 2},
]


def test_exporter_writes_valid_workbook(tmp_path):
    out = build_source_dictionary_workbook(
        OBJECTS, ATTRIBUTES,
        meta={"run_id": "run-export-test", "catalog": "insurance_source_discovery",
              "schema": "gw_pc_bronze", "scope_mode": "SCHEMA_ALL_TABLES"},
        out_path=tmp_path / "dd.xlsx")

    assert out.exists()
    wb = load_workbook(out)
    assert wb.sheetnames == ["Cover", "Objects", "Attributes"]
    # header + 1 object, header + 2 attributes
    assert wb["Objects"].max_row == 2
    assert wb["Attributes"].max_row == 3
    # facts rendered correctly
    assert wb["Attributes"]["C2"].value == "pol_id"
    assert wb["Attributes"]["F2"].value == "PRIMARY_KEY"
    assert wb["Attributes"]["E2"].value == "NO"
